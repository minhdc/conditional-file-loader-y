import os
import xml.etree.ElementTree as ET
from programy.clients.events.console.client import ConsoleBotClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#get context
aiml_location = '../../../raw-aiml/'
xlsx_location = '../../../xlsx/'



for each_xlsx_file in os.listdir(xlsx_location):
    if not os.path.isdir(each_xlsx_file):
        print("working on ", each_xlsx_file)
        workbook = load_workbook(filename=os.path.join(xlsx_location,each_xlsx_file))
        print("sheet names : ",workbook.sheetnames)        
        #aiml = ET.Element("aiml")
        with open(os.path.join(aiml_location,each_xlsx_file.replace('xlsx','aiml').lower()),'w',encoding='utf-8') as aiml_file:       
            aiml_file.write('<?xml version="1.0" encoding="UTF-8"?>')
            aiml_file.write("\n")
            aiml_file.write('<aiml version="2.0">')
            aiml_file.write("\n")

            for each_sheet in workbook.sheetnames:
                worksheet = workbook[each_sheet]
                print(worksheet)
                for row in range(2,10000):                                   
                    if worksheet.cell(column=1,row=row).value is None:                        
                        break
                    if worksheet.cell(column=3,row=row).value is None:                    
                        aiml_file.write("<category>") 
                        aiml_file.write("\n")
                        aiml_file.write("<pattern>")        
                        aiml_file.write(' # '+str(worksheet.cell(column=1,row=row).value).upper().replace('>',' ').replace('<',' ').replace('?',' ').replace('&',' và ').lstrip().rstrip()+' # ')        
                        aiml_file.write("</pattern>")
                        aiml_file.write("\n")
                        aiml_file.write("<template>")                        
                        aiml_file.write(str(worksheet.cell(column=2,row=row).value).replace('>',' ').replace('<',' ').replace('&','và'))                        
                        aiml_file.write("</template>")
                        aiml_file.write("\n")
                        aiml_file.write("</category>") 
                        aiml_file.write("\n")
                    elif worksheet.cell(column=4,row=row).value is None:
                        #pattern = ET.SubElement(category,"pattern").text = worksheet.cell(column=2,row=row).value
                        #template = ET.SubElement(category,"template").text = worksheet.cell(column=3,row=row).value
                        aiml_file.write("<category>") 
                        aiml_file.write("\n")
                        aiml_file.write("<pattern>")                        
                        aiml_file.write(' # '+str(worksheet.cell(column=2,row=row).value).upper().replace('>',' ').replace('<',' ').replace('?',' ').replace('&',' và ').lstrip().rstrip()+' # ')                        
                        aiml_file.write("</pattern>")
                        aiml_file.write("\n")
                        aiml_file.write("<template>")                        
                        aiml_file.write(str(worksheet.cell(column=3,row=row).value).replace('>',' ').replace('<',' ').replace('&','và'))                        
                        aiml_file.write("</template>")
                        aiml_file.write("\n")
                        aiml_file.write("</category>")                    
                        aiml_file.write("\n")
                    aiml_file.write("\n")
            aiml_file.write("</aiml>")
            
                
    
    
                

